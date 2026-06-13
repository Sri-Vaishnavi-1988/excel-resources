# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── PALETTE ──────────────────────────────────────────────────
P = dict(
    bg     = "0F1117", surface = "1A1D27", border = "2E3250",
    purple = "6C63FF", teal    = "00D4AA", red     = "FF6B6B",
    yellow = "FFD93D", orange  = "FF9F43", muted   = "8B92B8",
    white  = "E8EAF6", green   = "00C875", blue    = "0099FF",
    sql    = "004D61", py      = "3D3000", tab     = "3D1F00",
    gov    = "1A0A3D", ai      = "002B1A",
)

def fill(h):   return PatternFill("solid", fgColor=h)
def side(c):   return Side(style="thin", color=c)
def bdr(c="2E3250"): s=side(c); return Border(left=s,right=s,top=s,bottom=s)
def aln(h="left",v="center",wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)

def fnt(bold=False, color="E8EAF6", size=10, underline=None, italic=False):
    return Font(bold=bold, color=color, size=size, name="Segoe UI",
                underline=underline, italic=italic)

def hlink_font(size=10):
    return Font(bold=True, color="00AEFF", size=size, name="Segoe UI",
                underline="single")

def _c(ws, r, c, v="", bg="1A1D27", fg="E8EAF6", bold=False,
       wrap=False, h="left", size=10, border=True, link=None, link_is_internal=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = fill(bg)
    if link:
        cell.hyperlink = link
        cell.font = hlink_font(size)
    else:
        cell.font = fnt(bold=bold, color=fg, size=size)
    cell.alignment = aln(h, wrap=wrap)
    if border:
        cell.border = bdr()
    return cell

def section_header(ws, r, text, bg, cols=5):
    ws.merge_cells(f"A{r}:{get_column_letter(cols)}{r}")
    c = ws.cell(row=r, column=1, value=f"  {text}")
    c.fill = fill(bg); c.font = fnt(bold=True, color="FFFFFF", size=11)
    c.alignment = aln("left", "center"); ws.row_dimensions[r].height = 26

def col_header(ws, r, headers, bg, fg="FFFFFF"):
    ws.row_dimensions[r].height = 22
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=i, value=h)
        c.fill = fill(bg); c.font = fnt(bold=True, color=fg, size=10)
        c.alignment = aln("center"); c.border = bdr()

def data_row(ws, r, vals, bg, fg="E8EAF6", height=52):
    ws.row_dimensions[r].height = height
    for i, v in enumerate(vals, 1):
        _c(ws, r, i, v, bg=bg, fg=fg, wrap=True, size=10)

def nav_back(ws, r, target_sheet):
    """← Back to Contents hyperlink"""
    ws.row_dimensions[r].height = 22
    target = f"#'📋 Contents'!A1"
    ws.merge_cells(f"A{r}:E{r}")
    c = ws.cell(row=r, column=1, value="← Back to Contents")
    c.hyperlink = target
    c.font = hlink_font(10)
    c.alignment = aln("left", "center")
    c.fill = fill("1A1D27")

def sheet_title(ws, r, title, bg):
    ws.row_dimensions[r].height = 44
    ws.merge_cells(f"A{r}:E{r}")
    c = ws.cell(row=r, column=1, value=title)
    c.fill = fill(bg)
    c.font = Font(bold=True, color="FFFFFF", size=20, name="Segoe UI")
    c.alignment = aln("center", "center")

def col_w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def mcq_block(ws, start_r, questions, bg_q="1A1D27", bg_a="0F1117"):
    r = start_r
    for q_text, options, answer, explanation in questions:
        ws.row_dimensions[r].height = 22
        ws.merge_cells(f"A{r}:E{r}")
        c = ws.cell(row=r, column=1, value=q_text)
        c.fill = fill(bg_q); c.font = fnt(bold=True, color="FFD93D", size=10)
        c.alignment = aln("left", "center", wrap=True); c.border = bdr()
        r += 1
        for opt in options:
            ws.row_dimensions[r].height = 18
            ws.merge_cells(f"A{r}:E{r}")
            c = ws.cell(row=r, column=1, value=f"    {opt}")
            c.fill = fill(bg_q); c.font = fnt(color="8B92B8", size=10)
            c.alignment = aln("left", "center"); c.border = bdr()
            r += 1
        # Answer
        ws.row_dimensions[r].height = 20
        ws.merge_cells(f"A{r}:E{r}")
        c = ws.cell(row=r, column=1, value=f"✅ Answer: {answer}")
        c.fill = fill("002B1A"); c.font = fnt(bold=True, color="00D4AA", size=10)
        c.alignment = aln("left", "center"); c.border = bdr()
        r += 1
        ws.row_dimensions[r].height = 40
        ws.merge_cells(f"A{r}:E{r}")
        c = ws.cell(row=r, column=1, value=f"💡 {explanation}")
        c.fill = fill("001A12"); c.font = fnt(color="00D4AA", size=9, italic=True)
        c.alignment = aln("left", "center", wrap=True); c.border = bdr()
        r += 1
        # spacer
        ws.row_dimensions[r].height = 6
        for col in range(1, 6): ws.cell(row=r, column=col).fill = fill("0F1117")
        r += 1
    return r

def theory_block(ws, start_r, questions, bg="1A1D27"):
    r = start_r
    for q, ans in questions:
        ws.row_dimensions[r].height = 22
        ws.merge_cells(f"A{r}:E{r}")
        c = ws.cell(row=r, column=1, value=q)
        c.fill = fill("2E3250"); c.font = fnt(bold=True, color="6C63FF", size=10)
        c.alignment = aln("left", "center", wrap=True); c.border = bdr()
        r += 1
        ws.row_dimensions[r].height = 72
        ws.merge_cells(f"A{r}:E{r}")
        c = ws.cell(row=r, column=1, value=ans)
        c.fill = fill(bg); c.font = fnt(color="E8EAF6", size=10)
        c.alignment = aln("left", "center", wrap=True); c.border = bdr()
        r += 1
        ws.row_dimensions[r].height = 6
        for col in range(1, 6): ws.cell(row=r, column=col).fill = fill("0F1117")
        r += 1
    return r

# ═══════════════════════════════════════════════════════════════════
# SHEET 1 — CONTENTS
# ═══════════════════════════════════════════════════════════════════
ws_c = wb.active
ws_c.title = "📋 Contents"
ws_c.sheet_view.showGridLines = False
ws_c.sheet_properties.tabColor = P["purple"]

ws_c.row_dimensions[1].height = 52
ws_c.merge_cells("A1:F1")
c = ws_c.cell(row=1, column=1, value="Interview Prep Workbook — Sri Vaishnavi Devarashetty")
c.fill = fill(P["purple"]); c.font = Font(bold=True, color="FFFFFF", size=20, name="Segoe UI")
c.alignment = aln("center", "center")

ws_c.row_dimensions[2].height = 26
ws_c.merge_cells("A2:F2")
c = ws_c.cell(row=2, column=1, value="Data Analytics Interview  ·  SQL  ·  Python  ·  Tableau  ·  Governance  ·  AI Tools")
c.fill = fill("1A1D27"); c.font = fnt(color=P["teal"], size=12)
c.alignment = aln("center", "center")

# Table header
col_header(ws_c, 3, ["#", "Sheet", "Topics Covered", "Key Skills", "Questions", "Go To"], P["purple"])

sheets_info = [
    ("1", "📋 Contents",         "Overview, navigation, how to use this workbook",                               "Navigation",                         "—",    "#'📋 Contents'!A1"),
    ("2", "🗄️ SQL",              "Basics, JOINs, Window Functions, CTEs, Snowflake optimisation",                "SQL, Snowflake, Query design",        "40+",  "#'🗄️ SQL'!A1"),
    ("3", "🐍 Python & Pandas",  "Data structures, pandas, cleaning, GroupBy, lambda, file I/O",                "Python, pandas, numpy",               "35+",  "#'🐍 Python'!A1"),
    ("4", "📊 Tableau",          "LOD, Context Filters, Table Calcs, Parameters, Performance",                  "Tableau, BI design",                  "25+",  "#'📊 Tableau'!A1"),
    ("5", "🛡️ Data Governance",  "Quality, lineage, MDM, GDPR, data catalog, stewardship",                      "Governance, compliance, quality",     "20+",  "#'🛡️ Governance'!A1"),
    ("6", "🤖 AI & Tools",       "AI workflow, GitHub Copilot, ChatGPT, Snowflake Cortex, agentic AI",          "AI, Copilot, Cortex, LLMs",           "20+",  "#'🤖 AI & Tools'!A1"),
]

row_bgs = ["1A1D27","1A1D27","212440","1A1D27","212440","1A1D27"]
for i, (num, name, topics, skills, qs, link) in enumerate(sheets_info):
    r = i + 4
    ws_c.row_dimensions[r].height = 28
    bg = row_bgs[i]
    _c(ws_c, r, 1, num,    bg=bg, h="center", bold=True, fg=P["yellow"])
    _c(ws_c, r, 2, name,   bg=bg, bold=True,  fg=P["teal"])
    _c(ws_c, r, 3, topics, bg=bg, fg=P["white"], wrap=True)
    _c(ws_c, r, 4, skills, bg=bg, fg=P["muted"])
    _c(ws_c, r, 5, qs,     bg=bg, h="center", bold=True, fg=P["yellow"])
    c = ws_c.cell(row=r, column=6, value="→ Open Sheet")
    c.hyperlink = link
    c.font = hlink_font(10)
    c.fill = fill(bg)
    c.alignment = aln("center")
    c.border = bdr()

# How to use
ws_c.row_dimensions[11].height = 10
ws_c.row_dimensions[12].height = 28
ws_c.merge_cells("A12:F12")
c = ws_c.cell(row=12, column=1, value="HOW TO USE THIS WORKBOOK")
c.fill = fill(P["teal"]); c.font = fnt(bold=True, color="0F1117", size=13)
c.alignment = aln("left", "center")

how_to = [
    ("📌 Structure",  "Each sheet has sections: Key Concepts → Real-World Examples → MCQ Questions → Theoretical Interview Questions"),
    ("🎯 Study Mode", "Work through Key Concepts first, then test yourself with MCQs before reading the answers"),
    ("💼 Interview",  "Theoretical questions mirror real interview questions — practice answering them out loud using the STAR method"),
    ("⭐ Priority",   "SQL and Tableau are tested in almost every analytics interview. Start there."),
    ("🔗 Quiz",       "Also take the live quiz at: sri-vaishnavi-1988.github.io/quiz-app/"),
]
for i, (label, text) in enumerate(how_to):
    r = 13 + i
    ws_c.row_dimensions[r].height = 36
    _c(ws_c, r, 1, label, bg="212440", bold=True, fg=P["purple"])
    ws_c.merge_cells(f"B{r}:F{r}")
    _c(ws_c, r, 2, text, bg="1A1D27", fg=P["white"], wrap=True)

col_w(ws_c, [5, 24, 46, 28, 12, 16])

# ═══════════════════════════════════════════════════════════════════
# SHEET 2 — SQL
# ═══════════════════════════════════════════════════════════════════
ws_sql = wb.create_sheet("🗄️ SQL")
ws_sql.sheet_view.showGridLines = False
ws_sql.sheet_properties.tabColor = P["teal"]

sheet_title(ws_sql, 1, "🗄️ SQL — Key Concepts, Window Functions, Snowflake & Interview Q", "004D61")
nav_back(ws_sql, 2, "📋 Contents")

# Section 1 — Key Concepts
section_header(ws_sql, 3, "SECTION 1 — KEY CONCEPTS & BASICS", "005F6B")
col_header(ws_sql, 4, ["Concept", "Description", "Syntax / Example", "When to Use", "Gotcha"], "004D61")

sql_concepts = [
    ("SELECT basics",    "Retrieve columns from a table",                                             "SELECT col1, col2 FROM table WHERE condition",                              "Always the starting point",                    "SELECT * is slow on wide tables"),
    ("WHERE vs HAVING",  "WHERE filters rows before GROUP BY; HAVING filters after aggregation",       "WHERE status='active' | HAVING COUNT(*)>5",                                "WHERE for row-level; HAVING for group-level",  "Cannot use aggregate functions in WHERE"),
    ("JOINs — INNER",   "Returns only matching rows from both tables",                                "SELECT * FROM a INNER JOIN b ON a.id=b.id",                                 "When you only want matched records",           "Drops unmatched rows silently"),
    ("JOINs — LEFT",    "All rows from left table + matches from right (NULL if no match)",            "SELECT * FROM a LEFT JOIN b ON a.id=b.id",                                  "When you need all records from one table",     "NULL on right side means no match found"),
    ("GROUP BY",         "Collapse rows into groups for aggregation",                                  "SELECT dept, COUNT(*) FROM emp GROUP BY dept",                              "Summarising data",                             "Every non-aggregate column must be in GROUP BY"),
    ("ORDER BY",         "Sort result set ascending or descending",                                    "ORDER BY score DESC, name ASC",                                             "Final sort before display",                    "Expensive on large datasets without LIMIT"),
    ("CTE (WITH clause)","Named temporary result set reusable in the same query",                      "WITH cte AS (SELECT ...) SELECT * FROM cte",                                "Complex multi-step queries; replaces nested subqueries", "Only valid within the single query it is defined in"),
    ("Subquery",         "Query nested inside another query",                                          "SELECT * FROM (SELECT ... ) sub WHERE sub.col > 5",                         "One-off transformations",                      "Runs repeatedly; CTEs are faster for repeated use"),
    ("DISTINCT",         "Remove duplicate rows from result",                                          "SELECT DISTINCT department FROM employees",                                  "Count unique values",                          "DISTINCT on multiple cols = distinct combinations"),
    ("NULLIF(x, 0)",    "Return NULL if x equals 0; prevents division by zero",                       "SUM(sales)/NULLIF(SUM(target),0)",                                          "KPI percentage calculations",                  "NULL propagates — use COALESCE to default"),
    ("COALESCE",         "Return first non-NULL value from a list",                                    "COALESCE(score, 0)",                                                        "Fill missing values with a default",           "Use instead of ISNULL for portability"),
    ("CASE WHEN",        "Conditional logic inside a SELECT or WHERE",                                 "CASE WHEN score>=80 THEN 'Pass' ELSE 'Fail' END",                           "Categorise values, create flags",              "Evaluated left-to-right; first match wins"),
]
row_bgs_sql = ["1A1D27","151E2E"]
for i, row in enumerate(sql_concepts):
    r = 5 + i
    ws_sql.row_dimensions[r].height = 44
    bg = row_bgs_sql[i % 2]
    for j, v in enumerate(row, 1):
        _c(ws_sql, r, j, v, bg=bg, wrap=True, size=10)

# Section 2 — Window Functions
r2 = 5 + len(sql_concepts)
section_header(ws_sql, r2, "SECTION 2 — WINDOW FUNCTIONS", "005F6B")
col_header(ws_sql, r2+1, ["Function", "Purpose", "Real-Time Example", "Key Clause", "Difference from GROUP BY"], "004D61")

wf_data = [
    ("ROW_NUMBER()",    "Unique sequential number per row",                         "Deduplicate — keep latest enrollment per learner",            "OVER(PARTITION BY learner_id ORDER BY date DESC)",                          "Does NOT collapse rows"),
    ("RANK()",          "Rank with gaps after ties",                                "Rank courses by completion rate — 1,1,3 for tied courses",    "OVER(PARTITION BY dept ORDER BY rate DESC)",                                "Tied rows share rank; next rank is skipped"),
    ("DENSE_RANK()",   "Rank without gaps after ties",                             "Learner leaderboard — no skipped ranks",                       "OVER(ORDER BY score DESC)",                                                 "1,1,2 — no gaps unlike RANK()"),
    ("LAG(col, n)",    "Value from n rows before current row",                     "MoM change: current completions vs last month",                "OVER(ORDER BY month)",                                                      "Use LAG(col,1,0) to default NULL to 0"),
    ("LEAD(col, n)",   "Value from n rows after current row",                      "Preview next quarter target vs current",                       "OVER(ORDER BY month)",                                                      "LEAD/LAG are opposite directions"),
    ("SUM() OVER()",   "Running total without collapsing rows",                    "Cumulative completions over the year per learner",             "OVER(PARTITION BY id ORDER BY date ROWS UNBOUNDED PRECEDING)",              "GROUP BY collapses; window keeps all rows"),
    ("AVG() OVER()",   "Rolling average per window",                               "4-week rolling average of weekly completions",                 "OVER(ORDER BY week ROWS BETWEEN 3 PRECEDING AND CURRENT ROW)",              "ROWS BETWEEN controls the window frame"),
    ("PERCENT_RANK()", "Percentile (0.0 to 1.0) within partition",                "Top 10% learners by assessment score",                         "OVER(PARTITION BY course ORDER BY score)",                                  "Multiply by 100 to get percentile number"),
    ("NTILE(n)",        "Divide rows into n equal buckets",                         "Split learners into quartiles for segmentation",              "OVER(ORDER BY score)",                                                      "Bucket sizes may differ if rows not divisible by n"),
    ("QUALIFY (Snowflake)","Filter window function results inline — no outer query","Top 3 learners per department without subquery",              "QUALIFY RANK() OVER(...)<=3",                                               "Snowflake/BigQuery only — not standard SQL"),
]
for i, row in enumerate(wf_data):
    r = r2 + 2 + i
    ws_sql.row_dimensions[r].height = 44
    bg = row_bgs_sql[i % 2]
    for j, v in enumerate(row, 1):
        _c(ws_sql, r, j, v, bg=bg, wrap=True, size=10)

# Section 3 — Snowflake
r3 = r2 + 2 + len(wf_data)
section_header(ws_sql, r3, "SECTION 3 — SNOWFLAKE-SPECIFIC FEATURES", "005F6B")
col_header(ws_sql, r3+1, ["Feature", "What It Does", "Example", "Use Case", "Notes"], "004D61")

sf_data = [
    ("CLONE",          "Zero-copy instant copy of table/schema/DB",                         "CREATE TABLE orders_qa CLONE orders",                                  "QA testing without storage cost",           "Storage cost only incurred on changes"),
    ("Time Travel",    "Query historical state of a table",                                  "SELECT * FROM orders AT(OFFSET=>-86400)",                              "Detect silent row drops in pipelines",      "Default 1 day; Enterprise = up to 90 days"),
    ("Clustering Keys","Define physical sort order to reduce micro-partition scans",          "ALTER TABLE orders CLUSTER BY(order_date)",                            "Large fact tables filtered by date range",  "Adds maintenance cost on writes"),
    ("Result Cache",   "Cache query results for up to 24 hours",                             "ALTER SESSION SET USE_CACHED_RESULT=FALSE to disable",                 "Disable for live KPI spot checks",          "Cache invalidated if underlying data changes"),
    ("Cortex AI",      "Natural language to SQL / LLM features inside Snowflake",            "SELECT SNOWFLAKE.CORTEX.COMPLETE('llama3-70b','summarize...')",        "Business users querying data in English",   "Requires Snowflake Enterprise or above"),
]
for i, row in enumerate(sf_data):
    r = r3 + 2 + i
    ws_sql.row_dimensions[r].height = 44
    bg = row_bgs_sql[i % 2]
    for j, v in enumerate(row, 1):
        _c(ws_sql, r, j, v, bg=bg, wrap=True, size=10)

# Section 4 — MCQ
r4 = r3 + 2 + len(sf_data)
section_header(ws_sql, r4, "SECTION 4 — MCQ QUESTIONS", "005F6B")
sql_mcqs = [
    ("Q1. What does DENSE_RANK() return for values [100, 100, 90]?",
     ["A) 1, 1, 3","B) 1, 1, 2","C) 1, 2, 3","D) 0, 0, 1"],
     "B) 1, 1, 2",
     "DENSE_RANK gives no gaps after ties. Both 100s get rank 1, then 90 gets rank 2."),
    ("Q2. Which JOIN returns all rows from the LEFT table even if there is no match in the RIGHT?",
     ["A) INNER JOIN","B) RIGHT JOIN","C) LEFT JOIN","D) CROSS JOIN"],
     "C) LEFT JOIN",
     "LEFT JOIN returns all rows from the left table and matching rows from the right. Non-matching right rows appear as NULL."),
    ("Q3. What is the correct SQL execution order?",
     ["A) SELECT → FROM → WHERE → GROUP BY","B) FROM → WHERE → GROUP BY → HAVING → SELECT → ORDER BY","C) WHERE → FROM → SELECT → GROUP BY","D) GROUP BY → HAVING → FROM → SELECT"],
     "B) FROM → WHERE → GROUP BY → HAVING → SELECT → ORDER BY",
     "SQL processes: FROM (tables), WHERE (row filter), GROUP BY (aggregate), HAVING (group filter), SELECT (columns), ORDER BY (sort)."),
    ("Q4. Which Snowflake feature lets you query data from 24 hours ago without a backup?",
     ["A) CLONE","B) Time Travel","C) Result Cache","D) Cortex AI"],
     "B) Time Travel",
     "Time Travel lets you query historical table state using AT(OFFSET=>-86400) for 24 hours ago."),
    ("Q5. What makes a WHERE condition non-sargable?",
     ["A) Using = operator","B) Filtering on a primary key","C) Applying a function to an indexed column","D) Using AND with multiple conditions"],
     "C) Applying a function to an indexed column",
     "WHERE YEAR(sale_date)=2025 is non-sargable — the function prevents index use. Use WHERE sale_date BETWEEN '2025-01-01' AND '2025-12-31' instead."),
    ("Q6. What does QUALIFY do in Snowflake?",
     ["A) Filters rows before GROUP BY","B) Filters window function results without a subquery","C) Defines clustering keys","D) Enables Time Travel"],
     "B) Filters window function results without a subquery",
     "QUALIFY is a Snowflake-native clause that filters based on window function output — like HAVING but for window functions."),
    ("Q7. Which is faster for checking if a related row exists?",
     ["A) IN with subquery","B) EXISTS","C) They are always identical","D) OUTER JOIN with NULL check"],
     "B) EXISTS",
     "EXISTS short-circuits — it stops scanning as soon as it finds the first match. IN evaluates the full subquery first."),
    ("Q8. What is a CTE?",
     ["A) A permanent table","B) A cached query across sessions","C) A named temporary result set valid within one query","D) A Snowflake-only feature"],
     "C) A named temporary result set valid within one query",
     "A CTE (WITH clause) creates a named temporary result set valid only for the query it is defined in. Improves readability and allows reuse within the same query."),
]
r4 = mcq_block(ws_sql, r4+1, sql_mcqs)

# Section 5 — Theory
section_header(ws_sql, r4, "SECTION 5 — THEORETICAL INTERVIEW QUESTIONS", "005F6B")
sql_theory = [
    ("T1. How do you optimise a slow SQL query?",
     "1) Check if filters are on indexed/clustered columns — avoid wrapping columns in functions (non-sargable). 2) Replace SELECT * with explicit column names. 3) Use CTEs to break complex logic into readable steps. 4) Add appropriate indexes on JOIN keys and WHERE columns. 5) Read the execution plan — look for table scans and nested loops on large tables. 6) In Snowflake: ensure clustering keys match common filter columns; use RESULT_SCAN for repeated queries."),
    ("T2. What is the difference between WHERE and HAVING?",
     "WHERE filters individual rows BEFORE GROUP BY runs — you cannot use aggregate functions in WHERE. HAVING filters groups AFTER GROUP BY — you can reference aggregates like COUNT(*), SUM(), AVG(). Example: WHERE status='active' filters before grouping; HAVING COUNT(*)>5 filters groups with more than 5 rows."),
    ("T3. Explain window functions and when you use them over GROUP BY.",
     "Window functions (RANK, SUM OVER, LAG, LEAD) compute across a set of rows related to the current row without collapsing the result. GROUP BY collapses all rows in a group into one. Use window functions when you need both the row-level detail AND a group-level metric in the same query — e.g., each learner's score alongside the department average."),
    ("T4. How do you handle duplicates in a large dataset?",
     "1) Identify the grain — what combination of columns should be unique? 2) Use ROW_NUMBER() OVER(PARTITION BY key_cols ORDER BY created_at DESC) to rank duplicates. 3) Keep rn=1 (most recent). 4) In Snowflake: use QUALIFY ROW_NUMBER() OVER(...) = 1 for inline deduplication. 5) Investigate root cause — duplicates usually indicate a pipeline issue upstream."),
    ("T5. What is a correlated subquery and when is it a problem?",
     "A correlated subquery references a column from the outer query, so it re-runs for every row in the outer query — O(n) executions. Problematic on large tables. Fix: rewrite as a JOIN or window function. Example: instead of WHERE score > (SELECT AVG(score) FROM t WHERE dept=outer.dept), use AVG(score) OVER(PARTITION BY dept)."),
]
theory_block(ws_sql, r4+1, sql_theory)
col_w(ws_sql, [22, 28, 40, 36, 34])

# ═══════════════════════════════════════════════════════════════════
# SHEET 3 — PYTHON
# ═══════════════════════════════════════════════════════════════════
ws_py = wb.create_sheet("🐍 Python")
ws_py.sheet_view.showGridLines = False
ws_py.sheet_properties.tabColor = P["yellow"]

sheet_title(ws_py, 1, "🐍 Python & Pandas — Data Cleaning, Analysis & Interview Q", "3D3000")
nav_back(ws_py, 2, "📋 Contents")

section_header(ws_py, 3, "SECTION 1 — PYTHON BASICS FOR DATA ANALYSTS", "4A3800")
col_header(ws_py, 4, ["Concept", "Description", "Example", "Common Use", "Watch Out"], "3D3000")

py_basics = [
    ("List comprehension", "Compact way to build a list with a condition",         "[x*2 for x in range(10) if x%2==0]",                         "Transform and filter in one line",              "Readability suffers beyond 2 conditions"),
    ("Dictionary",         "Key-value store; O(1) lookup",                         "d = {'SQL':90, 'Python':85}; d['SQL']",                       "Store config, mappings, counts",                "KeyError if key missing — use .get(key, default)"),
    ("Lambda function",    "Anonymous one-liner function",                         "fn = lambda x: x**2; fn(4) → 16",                            "Quick transforms in apply()",                   "Avoid for complex logic — use def instead"),
    ("enumerate()",        "Loop with index and value together",                    "for i, val in enumerate(lst): print(i, val)",                 "Building indexed structures from lists",         "enumerate starts at 0 by default"),
    ("zip()",              "Pair elements from two iterables",                      "for a, b in zip(names, scores): ...",                         "Combining two parallel lists",                  "Stops at the shorter iterable — use zip_longest if needed"),
    ("try / except",       "Catch and handle exceptions gracefully",                "try: df=pd.read_csv(f) except FileNotFoundError: log()",      "File I/O, API calls, type conversions",          "Do not use bare except — always specify exception type"),
    ("f-strings",          "Fast string interpolation",                             "f'Score: {score:.1f}%'",                                      "Building messages, log entries",                "Available Python 3.6+ only"),
]
for i, row in enumerate(py_basics):
    r = 5 + i
    ws_py.row_dimensions[r].height = 44
    bg = ["1A1D27","212440"][i%2]
    for j, v in enumerate(row,1): _c(ws_py, r, j, v, bg=bg, wrap=True, size=10)

r2 = 5 + len(py_basics)
section_header(ws_py, r2, "SECTION 2 — PANDAS CORE OPERATIONS", "4A3800")
col_header(ws_py, r2+1, ["Operation", "Method", "Real-Time Example", "Output", "Notes"], "3D3000")

pandas_ops = [
    ("Read data",         "pd.read_csv(f, parse_dates=['date'])",                              "Load enrollment CSV from Snowflake export",              "DataFrame",                    "parse_dates avoids manual datetime conversion"),
    ("Inspect",           "df.shape, df.dtypes, df.isnull().sum()",                            "Quick DQ check after loading",                           "Dimensions, types, null counts","Always run after loading to catch issues early"),
    ("Filter rows",       "df[df['status']=='completed']",                                     "Get only completed enrollments",                         "Filtered DataFrame",           "Chain: df[(df.a>0) & (df.b<10)]"),
    ("Select columns",    "df[['col1','col2']] or df.col1",                                    "Extract learner_id and score",                           "DataFrame or Series",          "Double bracket = DataFrame; single = Series"),
    ("GroupBy + agg",     "df.groupby('dept').agg({'score':'mean','id':'count'})",              "Avg score and headcount per department",                 "Summary DataFrame",            "reset_index() to flatten MultiIndex"),
    ("transform()",       "df.groupby('dept')['score'].transform('mean')",                     "Add dept avg as new column alongside each row",          "Series aligned to original DF","Different from agg — keeps all rows"),
    ("apply()",           "df['col'].apply(lambda x: ...)",                                    "Flag high-risk learners based on days inactive",         "Series or DataFrame",          "Slower than vectorised ops — use only when needed"),
    ("merge()",           "df1.merge(df2, on='id', how='left')",                               "Join enrollment data to learner dimension table",        "Merged DataFrame",             "Specify how: inner/left/right/outer"),
    ("drop_duplicates()", "df.drop_duplicates(subset=['learner_id','course_id'], keep='last')","Remove duplicate enrollments keeping latest",           "Deduplicated DataFrame",       "Always specify subset — default checks all columns"),
    ("fillna / where",    "df['score'].where(df['status']=='completed', other=float('nan'))",  "Fill score with NaN for non-completed rows",             "Series",                       "Never fill missing scores with 0 — skews averages"),
    ("sort_values()",     "df.sort_values('date', ascending=False)",                           "Latest enrollments first",                               "Sorted DataFrame",             "inplace=True modifies in place — use carefully"),
    ("pivot_table()",     "df.pivot_table(values='score',index='dept',columns='course',aggfunc='mean')","Cross-tab: dept vs course avg scores",          "Matrix DataFrame",             "fill_value=0 to replace NaN in empty cells"),
    ("read_csv chunksize","pd.read_csv(f, chunksize=100_000)",                                 "Process 10GB file that won't fit in RAM",                "Iterator of DataFrames",       "Process each chunk and append results"),
    ("value_counts()",    "df['status'].value_counts(normalize=True)",                         "Completion rate distribution",                           "Series of counts/proportions", "normalize=True gives percentages"),
]
for i, row in enumerate(pandas_ops):
    r = r2 + 2 + i
    ws_py.row_dimensions[r].height = 44
    bg = ["1A1D27","212440"][i%2]
    for j, v in enumerate(row,1): _c(ws_py, r, j, v, bg=bg, wrap=True, size=10)

r3 = r2 + 2 + len(pandas_ops)
section_header(ws_py, r3, "SECTION 3 — DATA CLEANING PATTERNS", "4A3800")
col_header(ws_py, r3+1, ["Problem", "Detection", "Fix", "Example", "Notes"], "3D3000")

cleaning = [
    ("Missing values",      "df.isnull().sum()",                           "fillna, dropna, or flag as unknown",                "df['dept'].fillna('Unknown')",                    "Never use 0 for numeric KPIs — use NaN"),
    ("Duplicates",          "df.duplicated().sum()",                       "drop_duplicates(subset, keep='last')",               "df.drop_duplicates(subset=['id','date'])",        "Investigate root cause — likely a pipeline issue"),
    ("Wrong data types",    "df.dtypes",                                   "pd.to_datetime(), astype()",                         "df['date']=pd.to_datetime(df['date'])",           "Check dtypes immediately after loading"),
    ("Outliers (IQR)",      "q1,q3=df.quantile([.25,.75]); iqr=q3-q1",   "Flag or cap beyond Q1-1.5*IQR and Q3+1.5*IQR",     "df[(df.score>=lower)&(df.score<=upper)]",         "IQR is robust to skewed data; std dev is not"),
    ("Inconsistent strings","df['status'].unique()",                       "str.strip().str.lower().map(std_map)",               "df['status']=df['status'].str.lower().map(mapper)","Build a mapping dict for controlled vocabulary"),
    ("Negative durations",  "(df.end-df.start).dt.days < 0",              "Flag rows; investigate source",                      "df['dq_flag']='negative_duration'",               "Never silently drop — always flag first"),
]
for i, row in enumerate(cleaning):
    r = r3 + 2 + i
    ws_py.row_dimensions[r].height = 44
    bg = ["1A1D27","212440"][i%2]
    for j, v in enumerate(row,1): _c(ws_py, r, j, v, bg=bg, wrap=True, size=10)

r4 = r3 + 2 + len(cleaning)
section_header(ws_py, r4, "SECTION 4 — MCQ QUESTIONS", "4A3800")
py_mcqs = [
    ("Q1. What does df.groupby('dept')['score'].transform('mean') return?",
     ["A) One row per dept with mean score","B) A Series with original index, each row replaced by its group mean","C) Modifies the original DataFrame in place","D) Same as .agg('mean')"],
     "B) A Series with original index, each row replaced by its group mean",
     "transform() returns a result with the same index as the original DataFrame — every row gets its group's mean. Use this to add a group metric as a new column."),
    ("Q2. What is the difference between loc and iloc?",
     ["A) loc uses integer positions; iloc uses labels","B) loc uses labels/index names; iloc uses integer positions","C) They are identical in all cases","D) iloc is always faster than loc"],
     "B) loc uses labels/index names; iloc uses integer positions",
     "df.loc['row_label','col_name'] uses axis labels. df.iloc[0,1] uses 0-based integer positions. After reset_index they may match; after filtering they will not."),
    ("Q3. Which is the correct way to remove duplicate rows keeping the latest date?",
     ["A) df.remove_duplicates()","B) df.drop_duplicates(subset=['id'], keep='last') after sorting by date","C) df.unique()","D) df.deduplicate(keep='last')"],
     "B) df.drop_duplicates(subset=['id'], keep='last') after sorting by date",
     "First sort by date ascending, then drop_duplicates with keep='last' keeps the most recent row. Or sort descending and use keep='first'."),
    ("Q4. How do you process a 10GB CSV file that won't fit in memory?",
     ["A) Increase RAM","B) pd.read_csv with chunksize parameter","C) Convert to JSON first","D) pandas cannot handle files over 2GB"],
     "B) pd.read_csv with chunksize parameter",
     "pd.read_csv('file.csv', chunksize=100_000) returns an iterator of DataFrames. Process and aggregate each chunk, then combine results."),
    ("Q5. What does .apply() do in pandas?",
     ["A) Applies a function element-wise to a DataFrame only","B) Applies a function along an axis of a DataFrame or Series","C) Always faster than vectorised operations","D) Only works on numeric columns"],
     "B) Applies a function along an axis of a DataFrame or Series",
     "apply() works on a Series (element-wise) or DataFrame (row/column-wise via axis param). It is generally slower than vectorised NumPy operations — use only when no vectorised alternative exists."),
    ("Q6. What is the IQR method for outlier detection?",
     ["A) Remove values more than 2 standard deviations from mean","B) Flag values below Q1-1.5*IQR or above Q3+1.5*IQR","C) Remove the top and bottom 5% of values","D) Use median absolute deviation only"],
     "B) Flag values below Q1-1.5*IQR or above Q3+1.5*IQR",
     "IQR = Q3 - Q1. Values outside [Q1 - 1.5*IQR, Q3 + 1.5*IQR] are outliers. This method is robust to skewed distributions unlike the z-score method."),
]
r4 = mcq_block(ws_py, r4+1, py_mcqs)

section_header(ws_py, r4, "SECTION 5 — THEORETICAL INTERVIEW QUESTIONS", "4A3800")
py_theory = [
    ("T1. What is the difference between apply(), map(), and transform() in pandas?",
     "map() is element-wise on a Series — maps each value through a function or dict. apply() works on a Series element-wise OR a DataFrame row/column-wise — more flexible, slower. transform() returns a result aligned to the original index — use with groupby to add group stats back to each row. Rule: use vectorised ops first, then map, then apply, then transform."),
    ("T2. How do you handle missing values in a key ID column?",
     "Never impute key identifiers (learner_id, order_id) — missing IDs indicate a pipeline issue. Steps: 1) Quantify: how many rows have missing IDs? 2) Investigate: is it a join failure or a source system issue? 3) Flag with a DQ alert — do not silently drop or fill. 4) Escalate to data engineering if systematic. 5) Document the decision in your DQ log."),
    ("T3. When would you use Dask or Polars instead of pandas?",
     "Dask: when data exceeds RAM and you want pandas-like syntax with parallel processing across chunks. Polars: when you need 5-20x faster performance than pandas on in-memory data (uses Rust, lazy evaluation, SIMD). pandas is still best for interactive EDA on datasets under ~1GB. In analytics at scale: push heavy processing to Snowflake SQL before loading to Python."),
    ("T4. Explain the 80/20 rule in data science projects.",
     "Approximately 80% of project time is spent on data preparation — cleaning, joining, validating, transforming. Only 20% is on analysis and modelling. This is why strong pandas and SQL skills matter more than ML algorithms for most analytics roles. Automation (Python scripts for ETL) is how analysts reclaim that 80%."),
]
theory_block(ws_py, r4+1, py_theory)
col_w(ws_py, [22, 32, 40, 28, 34])

# ═══════════════════════════════════════════════════════════════════
# SHEET 4 — TABLEAU
# ═══════════════════════════════════════════════════════════════════
ws_tab = wb.create_sheet("📊 Tableau")
ws_tab.sheet_view.showGridLines = False
ws_tab.sheet_properties.tabColor = P["orange"]

sheet_title(ws_tab, 1, "📊 Tableau — LOD, Filters, Performance & Interview Q", "3D1F00")
nav_back(ws_tab, 2, "📋 Contents")

section_header(ws_tab, 3, "SECTION 1 — TABLEAU KEY CONCEPTS", "4A2800")
col_header(ws_tab, 4, ["Concept", "Description", "Real Example", "Common Mistake", "Pro Tip"], "3D1F00")

tab_concepts = [
    ("FIXED LOD",      "Computes at a fixed dimension, ignores view filters",                                     "{FIXED [Dept]: AVG([Completion Rate])}",                     "Forgetting it ignores dimension filters",           "Use context filter to force FIXED to respect a filter"),
    ("INCLUDE LOD",    "Adds extra granularity below the view level",                                             "{INCLUDE [Course]: AVG([Score])}",                           "Using when FIXED is actually needed",               "Prevents Simpson's Paradox in aggregated views"),
    ("EXCLUDE LOD",    "Removes a dimension from the calculation",                                                "SUM([Sales])/{EXCLUDE [Product]:SUM([Sales])}",              "Applying to wrong level",                           "Perfect for % of total without table calcs"),
    ("Context Filter", "Executes first; defines the data pool for all other filters",                             "Force Top N to work correctly alongside a region filter",    "Not setting context = Top N sees unfiltered data",  "Right-click filter → Add to Context"),
    ("Table Calculation","Runs in Tableau after data is retrieved",                                               "RUNNING_SUM, WINDOW_AVG, RANK, PERCENT_TOTAL",               "Using as a filter the same as dimensions",          "Compute using: Table Down/Across/Cell"),
    ("Parameters",     "User-controlled input driving calcs, filters, ref lines",                                 "Top N slider (1-20), date range selector",                   "Forgetting to wire parameter to a calculated field","Parameters don't filter data — they feed calculated fields"),
    ("Calculated Field","Custom metric or dimension in Tableau",                                                  "IF [Score]>=80 THEN 'Pass' ELSE 'Fail' END",                 "Complex calc slowing down large extracts",          "Pre-aggregate in SQL when possible"),
    ("Extract vs Live","Extract = local Hyper file; Live = direct DB query",                                      "Extract for dashboards with millions of rows",               "Using live on large tables = slow dashboard",       "Schedule extract refreshes during off-peak hours"),
    ("Row-Level Security","Show only data relevant to the logged-in user",                                        "Manager sees only their department's learners",              "Hardcoding user names — use USERNAME() function",  "Use User Filter + entitlement table for scalability"),
    ("Blending vs Join","Join = combine before aggregation; Blend = combine after",                               "Blend: secondary data source for targets vs actuals",        "Blend loses row-level granularity",                 "Prefer joins (native or SQL) over blending for performance"),
]
for i, row in enumerate(tab_concepts):
    r = 5 + i
    ws_tab.row_dimensions[r].height = 44
    bg = ["1A1D27","212440"][i%2]
    for j, v in enumerate(row,1): _c(ws_tab, r, j, v, bg=bg, wrap=True, size=10)

r2 = 5 + len(tab_concepts)
section_header(ws_tab, r2, "SECTION 2 — PERFORMANCE OPTIMISATION", "4A2800")
col_header(ws_tab, r2+1, ["Problem", "Symptom", "Fix", "Example", "Impact"], "3D1F00")

perf_data = [
    ("Too many marks",        "Dashboard slow to render",              "Reduce view marks — aggregate more",                       "Show weekly instead of daily granularity",                          "High"),
    ("Live on large DB",      "Every interaction triggers a DB query", "Switch to extract for static/daily-refresh data",           "Daily L&D completion extract",                                      "High"),
    ("LOD on high cardinality","Extract build slow",                   "Pre-aggregate in Snowflake SQL instead",                    "Move FIXED to a Snowflake CTE",                                     "Medium"),
    ("Wrong context filter",  "Top N shows wrong results",             "Add dimension filter to context first",                     "Right-click region filter → Add to Context",                        "Medium"),
    ("Complex calcs in viz",  "Slow per-interaction",                  "Move calculations to data source (custom SQL)",             "Pre-compute risk flag in Snowflake",                                "Medium"),
]
for i, row in enumerate(perf_data):
    r = r2 + 2 + i
    ws_tab.row_dimensions[r].height = 44
    bg = ["1A1D27","212440"][i%2]
    for j, v in enumerate(row,1): _c(ws_tab, r, j, v, bg=bg, wrap=True, size=10)

r3 = r2 + 2 + len(perf_data)
section_header(ws_tab, r3, "SECTION 3 — MCQ QUESTIONS", "4A2800")
tab_mcqs = [
    ("Q1. Which LOD expression type ignores all view-level dimension filters?",
     ["A) INCLUDE","B) EXCLUDE","C) FIXED","D) TOTAL"],
     "C) FIXED",
     "FIXED computes at the dimension(s) you specify, completely ignoring what is in the view. Use a context filter if you need FIXED to respect a specific filter."),
    ("Q2. What is the difference between a regular filter and a context filter?",
     ["A) They are functionally identical","B) Regular filters apply before context filters","C) Context filters execute first and create a temporary table that all other filters operate on","D) Context filters only apply to measures"],
     "C) Context filters execute first and create a temporary table that all other filters operate on",
     "Context filters run first and define the data pool for the rest of the view. Other filters, FIXED LOD expressions, and Top N filters operate on this reduced dataset."),
    ("Q3. When do table calculations compute?",
     ["A) At the data source before retrieval","B) During extract refresh","C) After data is retrieved, at the visualisation level in Tableau","D) Before aggregation in the database"],
     "C) After data is retrieved, at the visualisation level in Tableau",
     "Table calculations (RUNNING_SUM, WINDOW_AVG, RANK) run entirely inside Tableau after the data is fetched. They cannot be pushed to the database."),
    ("Q4. What are Parameters in Tableau used for?",
     ["A) Filtering data at the database level","B) Creating dynamic user-controlled inputs that drive calculations, filters, and reference lines","C) Connecting to multiple data sources","D) Scheduling dashboard refreshes"],
     "B) Creating dynamic user-controlled inputs that drive calculations, filters, and reference lines",
     "Parameters let users input values (slider, dropdown, text) that calculated fields and filters can reference. They do not directly filter data — you must use them in a calculated field or filter condition."),
    ("Q5. Best practice for showing each course row alongside its department average?",
     ["A) Use WINDOW_AVG table calculation","B) Use FIXED LOD: {FIXED [Department]: AVG([Score])}","C) Use a dual-axis chart","D) Use blending with a secondary data source"],
     "B) Use FIXED LOD: {FIXED [Department]: AVG([Score])}",
     "FIXED LOD computes the department average independently of what is in the view. This lets you show each course row (view granularity = course) alongside the dept average in the same row."),
]
r3 = mcq_block(ws_tab, r3+1, tab_mcqs)

section_header(ws_tab, r3, "SECTION 4 — THEORETICAL INTERVIEW QUESTIONS", "4A2800")
tab_theory = [
    ("T1. Explain the difference between FIXED, INCLUDE, and EXCLUDE LOD expressions.",
     "FIXED: computes at specified dimension(s) regardless of view — ignores dimension/measure filters (only respects context filters). Best for cohort metrics, customer LTV, baseline benchmarks.\nINCLUDE: adds a dimension to the view's granularity — higher detail than the view. Best for sub-group averages.\nEXCLUDE: removes a dimension from the view — lower detail. Best for % of total, grand totals in a detailed view."),
    ("T2. How do you make a Tableau dashboard faster?",
     "1) Pre-aggregate in Snowflake SQL — push heavy computation to the warehouse. 2) Switch from live to extract for non-real-time dashboards. 3) Reduce marks — show weekly instead of daily. 4) Use context filters for large dimension filters. 5) Move complex LOD calculations to SQL CTEs. 6) Use Performance Recorder (Help menu) to identify the slowest queries."),
    ("T3. Why would a Top N filter show incorrect results, and how do you fix it?",
     "Top N filters operate on the data the view can see. If a dimension filter (e.g., region) is applied after the Top N is evaluated, the Top N sees the full dataset, not the filtered one. Fix: add the dimension filter to Context (right-click → Add to Context). This forces the region filter to execute first, then Top N operates on the reduced dataset."),
]
theory_block(ws_tab, r3+1, tab_theory)
col_w(ws_tab, [22, 28, 40, 34, 34])

# ═══════════════════════════════════════════════════════════════════
# SHEET 5 — GOVERNANCE
# ═══════════════════════════════════════════════════════════════════
ws_gov = wb.create_sheet("🛡️ Governance")
ws_gov.sheet_view.showGridLines = False
ws_gov.sheet_properties.tabColor = P["purple"]

sheet_title(ws_gov, 1, "🛡️ Data Governance — Quality, Lineage, GDPR & Interview Q", "1A0A3D")
nav_back(ws_gov, 2, "📋 Contents")

section_header(ws_gov, 3, "SECTION 1 — KEY CONCEPTS", "23154A")
col_header(ws_gov, 4, ["Concept", "Definition", "Real Example", "Why It Matters", "Tool/Framework"], "1A0A3D")

gov_concepts = [
    ("Data Quality",         "Accuracy, completeness, consistency, timeliness of data",                       "Row count reconciliation before weekly briefing",                   "Wrong data destroys stakeholder trust",            "Great Expectations, dbt tests, custom SQL checks"),
    ("Data Lineage",         "Trail showing where data originated, how it was transformed, and where it flows","raw_events → Snowflake staging → fact table → Tableau dashboard",   "Impact analysis: what breaks if this table changes?","dbt, Collibra, DataHub, Microsoft Purview"),
    ("Data Catalog",         "Centralised inventory of data assets with metadata, owners, quality scores",     "Searchable index of all Snowflake tables at Apple",                 "Analysts find and trust data faster",              "Alation, Collibra, DataHub, Snowflake Data Marketplace"),
    ("Data Dictionary",      "Document defining each field: name, type, description, owner, valid values",     "KPI definitions shared across NVIDIA analytics team",              "Prevents conflicting KPI definitions",             "Confluence, SharePoint, dbt docs"),
    ("MDM",                  "Single authoritative golden record for key business entities across systems",    "One customer_id across CRM, ERP, and support systems",             "Eliminates same customer — three IDs problem",     "SAP MDG, Informatica, Talend MDM"),
    ("Data Steward",         "Person responsible for day-to-day quality management within a domain",           "Analytics lead defining KPI rules and resolving quality issues",    "Bridges business owners and IT custodians",        "Role, not a tool — defined in governance policy"),
    ("Data Owner",           "Executive accountable for an entire data domain",                                "Head of L&D owns all learning platform data",                       "Accountable for access, quality, and compliance",  "Role defined in data governance charter"),
    ("GDPR Storage Limitation","Personal data retained only as long as necessary for stated purpose",          "Learner activity logs purged after 3 years per policy",             "Legal compliance — fines up to 4% of global turnover","Microsoft Purview, OneTrust, custom retention policies"),
    ("PII",                  "Data that can identify an individual: name, email, ID, IP address",              "Masking learner names in dashboard exports",                        "Privacy compliance, data breach risk reduction",   "Snowflake Dynamic Data Masking, column-level security"),
    ("Data Mesh",            "Decentralised architecture where domain teams own their data products",           "L&D team owns their data product; Finance owns theirs",             "Scales governance without central bottleneck",     "Architectural pattern — enabled by dbt, Databricks, Snowflake"),
]
for i, row in enumerate(gov_concepts):
    r = 5 + i
    ws_gov.row_dimensions[r].height = 44
    bg = ["1A1D27","212440"][i%2]
    for j, v in enumerate(row,1): _c(ws_gov, r, j, v, bg=bg, wrap=True, size=10)

r2 = 5 + len(gov_concepts)
section_header(ws_gov, r2, "SECTION 2 — PRE-PUBLISH CHECKLIST (5 SECTIONS)", "23154A")
col_header(ws_gov, r2+1, ["Section", "Check", "How to Verify", "Failure Action", "Priority"], "1A0A3D")

checklist = [
    ("Source Validation","Row count within ±5% of yesterday",                   "SELECT COUNT(*) vs Time Travel count",                    "Hold report; investigate pipeline",               "P1 🔴"),
    ("Source Validation","No NULLs in primary key columns",                      "SELECT COUNT(*) WHERE id IS NULL",                        "Fix at source; flag in DQ log",                   "P1 🔴"),
    ("KPI Accuracy",     "Totals cross-check against source system",             "Manual spot check on 3 departments",                      "Do not publish; fix calculation",                 "P1 🔴"),
    ("KPI Accuracy",     "Percentages sum to 100% where expected",               "Visual check + SUM() validation query",                   "Review denominator — NULLIF issue?",              "P1 🔴"),
    ("Tableau Checks",   "All filters apply correctly to all sheets",            "Click each filter and observe all sheets",                 "Fix filter scope or add to context",              "P2 🟡"),
    ("Tableau Checks",   "LOD expressions validated against raw SQL",            "Run equivalent SQL; compare totals",                      "Rebuild LOD with correct FIXED dimension",         "P1 🔴"),
    ("Compliance",       "PII fields masked or excluded",                        "Check each exported field against PII register",          "Remove field; escalate to data owner",            "P1 🔴"),
    ("Sign-off",         "Stakeholder sign-off received",                        "Email or Jira ticket approval",                           "Do not go live without approval",                 "P2 🟡"),
]
for i, row in enumerate(checklist):
    r = r2 + 2 + i
    ws_gov.row_dimensions[r].height = 36
    bg = ["1A1D27","212440"][i%2]
    for j, v in enumerate(row,1): _c(ws_gov, r, j, v, bg=bg, wrap=True, size=10)

r3 = r2 + 2 + len(checklist)
section_header(ws_gov, r3, "SECTION 3 — MCQ QUESTIONS", "23154A")
gov_mcqs = [
    ("Q1. What is a data catalog?",
     ["A) A database of dictionaries for NLP models","B) A centralised inventory of data assets with metadata, owners, quality scores","C) A tool for encrypting sensitive data","D) A list of all governance policies"],
     "B) A centralised inventory of data assets with metadata, owners, quality scores",
     "Data catalogs (Alation, Collibra, DataHub) are searchable inventories where analysts find data assets, understand their lineage, and assess quality before using them."),
    ("Q2. Under GDPR, how long may personal data be retained?",
     ["A) Maximum 5 years","B) Maximum 10 years","C) Only as long as necessary for the stated purpose","D) Indefinitely if encrypted at rest"],
     "C) Only as long as necessary for the stated purpose",
     "GDPR's storage limitation principle requires personal data be kept no longer than necessary. There is no fixed maximum — it depends on the stated purpose and legal basis."),
    ("Q3. What is Master Data Management (MDM)?",
     ["A) Managing the master branch of a data repo","B) Creating a single authoritative golden record for key business entities across systems","C) A methodology for big data at petabyte scale","D) Archiving historical data to cold storage"],
     "B) Creating a single authoritative golden record for key business entities across systems",
     "MDM de-duplicates and reconciles records for key entities like customers, products, and suppliers across multiple source systems to create one trusted golden record."),
    ("Q4. What is the role of a Data Steward?",
     ["A) Executive accountable for the entire data domain","B) IT person responsible for physical storage","C) Person responsible for day-to-day data quality management within a domain","D) The analyst who queries and reports on data"],
     "C) Person responsible for day-to-day data quality management within a domain",
     "Data Stewards own operational data quality: defining business rules, resolving quality issues, and acting as the bridge between data owners (executives) and IT custodians (engineers)."),
]
r3 = mcq_block(ws_gov, r3+1, gov_mcqs)

section_header(ws_gov, r3, "SECTION 4 — THEORETICAL INTERVIEW QUESTIONS", "23154A")
gov_theory = [
    ("T1. How do you ensure data quality before publishing a dashboard?",
     "I run a 5-section pre-publish checklist: 1) Source validation — row count within ±5% of yesterday using Snowflake Time Travel. 2) KPI accuracy — cross-check totals against source system for 3 sample departments. 3) Tableau filter validation — click every filter and confirm all sheets update correctly. 4) LOD validation — run equivalent SQL and compare totals. 5) Compliance check — confirm no PII fields in exports. Only publish after all P1 checks pass."),
    ("T2. What is data lineage and why does it matter?",
     "Data lineage is the documented trail of where data originated, how it was transformed at each step, and where it flows downstream. It matters because: 1) Impact analysis — if the source table changes, you know immediately which dashboards break. 2) Debugging — when a KPI looks wrong, you can trace back through every transformation. 3) Compliance — GDPR and audit requirements often mandate knowing exactly where personal data flows. Tools: dbt docs auto-generate lineage graphs; Collibra and DataHub provide enterprise-grade lineage."),
    ("T3. How would you handle a situation where two teams have different definitions of the same KPI?",
     "This is a governance problem, not a technical one. Steps: 1) Facilitate a KPI definition workshop — bring both teams together with the data owner. 2) Document both definitions and understand the business reason for each. 3) Determine if one definition is correct or if both are valid for different use cases. 4) Agree on a canonical definition and document it in the data dictionary (dbt docs or Confluence). 5) Update all affected reports to use the agreed definition. 6) Communicate the change with a change log so stakeholders understand why numbers may differ from previous reports."),
]
theory_block(ws_gov, r3+1, gov_theory)
col_w(ws_gov, [22, 34, 34, 32, 30])

# ═══════════════════════════════════════════════════════════════════
# SHEET 6 — AI & TOOLS
# ═══════════════════════════════════════════════════════════════════
ws_ai = wb.create_sheet("🤖 AI & Tools")
ws_ai.sheet_view.showGridLines = False
ws_ai.sheet_properties.tabColor = P["green"]

sheet_title(ws_ai, 1, "🤖 AI & Tools — Daily Workflow, Agentic AI & Interview Q", "002B1A")
nav_back(ws_ai, 2, "📋 Contents")

section_header(ws_ai, 3, "SECTION 1 — AI TOOLS FOR ANALYTICS", "003D28")
col_header(ws_ai, 4, ["Tool", "Primary Use in Analytics", "Real Example", "Governance Rule", "Limitation"], "002B1A")

ai_tools = [
    ("GitHub Copilot",   "SQL and Python scaffolding inside code editor",                             "Scaffold Snowflake CTE + window function structure",              "Review every generated line before running — catch wrong PARTITION BY", "Does not know your schema — always validate column names"),
    ("ChatGPT",          "Stakeholder communications, documentation drafts, data dictionary drafts",  "Draft Monday L&D briefing from KPI CSV",                          "All numbers manually verified before sending",   "Hallucination risk — never trust numbers without verification"),
    ("Claude",           "Complex analytical reasoning, Tableau LOD logic, business logic review",    "Suggest FIXED LOD for a tricky aggregation edge case",            "Treat as a senior peer suggestion — not ground truth", "Context window limit on very large documents"),
    ("Snowflake Cortex", "Natural language to SQL; LLM calls inside Snowflake",                       "Business user asks top 5 courses by completion rate in plain English","Analytics team reviews results before sharing", "Requires Snowflake Enterprise; NL-to-SQL accuracy varies"),
    ("Tableau Pulse",    "AI-generated natural language insights on dashboard metrics",               "Auto-narrate why completion rate dropped 15% this week",          "Human review before insight reaches exec audience","In preview/limited availability; output quality varies"),
]
for i, row in enumerate(ai_tools):
    r = 5 + i
    ws_ai.row_dimensions[r].height = 44
    bg = ["1A1D27","212440"][i%2]
    for j, v in enumerate(row,1): _c(ws_ai, r, j, v, bg=bg, wrap=True, size=10)

r2 = 5 + len(ai_tools)
section_header(ws_ai, r2, "SECTION 2 — AI DAILY WORKFLOW", "003D28")
col_header(ws_ai, r2+1, ["Time of Day", "Task", "Tool Used", "Your Role (Human Step)", "Why It Matters"], "002B1A")

workflow = [
    ("Morning",     "Summarise overnight KPI alerts into 3-line stakeholder brief",     "Claude / ChatGPT",   "Edit tone; validate all numbers against Snowflake",                  "Saves 30 min; human ensures accuracy"),
    ("Mid-day",     "Scaffold Snowflake SQL — CTEs, window functions",                  "GitHub Copilot",     "Review every line; test on sample data; validate output",            "Catches wrong PARTITION BY or column name errors"),
    ("Mid-day",     "Suggest Tableau LOD expressions for complex aggregations",          "Claude",             "Validate in Tableau Desktop before publishing",                      "Surfaces approaches you might not have considered"),
    ("Mid-day",     "Generate first draft of data dictionary entries",                   "ChatGPT",            "Correct field definitions; check against source schema",             "Dictionary draft in 10 min vs 2 hours manually"),
    ("Afternoon",   "Translate analyst findings into business-language summary",         "Claude / ChatGPT",   "Edit every sentence; approve final copy before sending",             "Ensures C-suite language without hours of rewriting"),
    ("End of Day",  "Convert meeting notes into clean Jira tickets",                     "ChatGPT",            "Add acceptance criteria; assign correctly",                          "Consistent ticket quality; saves 20 min per sprint"),
]
for i, row in enumerate(workflow):
    r = r2 + 2 + i
    ws_ai.row_dimensions[r].height = 44
    bg = ["1A1D27","212440"][i%2]
    for j, v in enumerate(row,1): _c(ws_ai, r, j, v, bg=bg, wrap=True, size=10)

r3 = r2 + 2 + len(workflow)
section_header(ws_ai, r3, "SECTION 3 — AGENTIC AI CONCEPTS", "003D28")
col_header(ws_ai, r3+1, ["Concept", "Definition", "Analytics Example", "Current Maturity", "Your Position"], "002B1A")

agentic = [
    ("Agentic AI",             "AI that plans and takes multi-step actions autonomously",                  "Pull KPIs → detect anomaly → draft Slack alert → post to channel",  "Early production: Snowflake Cortex, n8n, LangGraph",  "Exploring — not yet in production"),
    ("RAG",                    "LLM answers questions by retrieving relevant documents first",              "Ask AI about a metric; it retrieves the data dictionary first",       "Production-ready in several tools",                   "Foundational concept to understand"),
    ("Tool Use / Function Call","LLM decides which tool/API to call based on user request",                "NL query → Cortex decides whether to run SQL or look up docs",        "Production in ChatGPT, Claude, Copilot",              "What makes Copilot useful in your workflow"),
    ("n8n / LangGraph",         "Visual (n8n) or Python (LangGraph) frameworks for agentic workflows",    "Weekly: pull CSV → summarise with AI → post to email",                "Growing adoption in analytics teams",                 "Explore as a weekend project — strong interview signal"),
]
for i, row in enumerate(agentic):
    r = r3 + 2 + i
    ws_ai.row_dimensions[r].height = 44
    bg = ["1A1D27","212440"][i%2]
    for j, v in enumerate(row,1): _c(ws_ai, r, j, v, bg=bg, wrap=True, size=10)

r4 = r3 + 2 + len(agentic)
section_header(ws_ai, r4, "SECTION 4 — MCQ QUESTIONS", "003D28")
ai_mcqs = [
    ("Q1. What is the most important governance rule when using AI tools in analytics?",
     ["A) Use AI for all SQL — it is always more accurate","B) Every AI output — numbers, queries, messages — gets human review before it reaches a stakeholder","C) Only use AI for documentation, never for SQL","D) AI tools should replace junior analysts entirely"],
     "B) Every AI output — numbers, queries, messages — gets human review before it reaches a stakeholder",
     "AI tools can hallucinate numbers, use wrong column names, and miss business context. The analyst is accountable for every output. Human-in-the-loop review is non-negotiable."),
    ("Q2. What is agentic AI?",
     ["A) AI that answers single prompts very accurately","B) AI that can plan and take multi-step actions autonomously","C) AI that only works inside Snowflake","D) AI that replaces the need for SQL"],
     "B) AI that can plan and take multi-step actions autonomously",
     "Agentic AI goes beyond single prompt-response — it can plan a sequence of actions, call tools (APIs, databases), and complete multi-step tasks without a human prompt at each step."),
    ("Q3. What is Snowflake Cortex primarily used for in analytics?",
     ["A) Replacing Tableau dashboards","B) Natural language to SQL and LLM capabilities inside Snowflake","C) Automated data cleaning only","D) Replacing Python pipelines"],
     "B) Natural language to SQL and LLM capabilities inside Snowflake",
     "Snowflake Cortex enables business users to query data in plain English (natural language to SQL) and provides LLM functions like COMPLETE, SENTIMENT, and TRANSLATE directly in SQL."),
    ("Q4. Which GitHub Copilot governance step is most critical for analytics professionals?",
     ["A) Accept all suggestions immediately for speed","B) Never use Copilot for SQL","C) Review every generated line before running and validate output against sample data","D) Only use Copilot for Python, not SQL"],
     "C) Review every generated line before running and validate output against sample data",
     "Copilot does not know your schema or business rules. It can generate syntactically correct SQL that is logically wrong — wrong JOIN type, wrong PARTITION BY, wrong column name. Always validate."),
]
r4 = mcq_block(ws_ai, r4+1, ai_mcqs)

section_header(ws_ai, r4, "SECTION 5 — THEORETICAL INTERVIEW QUESTIONS", "003D28")
ai_theory = [
    ("T1. How do you use AI in your day-to-day analytics work? (The perfect answer)",
     "1) The tool: I use ChatGPT, GitHub Copilot, and Claude. 2) The use case: primarily for scaffolding SQL (CTEs, window functions), drafting stakeholder communications, and suggesting Tableau LOD approaches. 3) The governance rule: every AI output — every query, every message — gets human review before it reaches a stakeholder or runs in production. 4) The mindset: I treat AI as a senior peer who is very fast but sometimes wrong. I bring the business context and domain knowledge; AI brings speed and breadth."),
    ("T2. What is agentic AI and how might it apply to analytics?",
     "Agentic AI is when an AI system can plan and take multi-step actions autonomously — calling tools, APIs, or databases without a human prompt at each step. In analytics, a near-future example: the agent detects a KPI anomaly at 2am → queries Snowflake for root cause → drafts a Slack alert → posts to the on-call channel. Today this requires n8n or LangGraph wiring. I am exploring this as a weekend project with n8n connecting to Snowflake. The analyst's role shifts from building the workflow to defining the business rules and reviewing outputs."),
    ("T3. What are the risks of using AI tools in analytics, and how do you mitigate them?",
     "Risks: 1) Hallucination — AI generates plausible but wrong numbers or SQL. Mitigation: always validate queries on sample data; never share AI output with stakeholders without manual number check. 2) Schema ignorance — Copilot doesn't know your column names or business logic. Mitigation: provide context in the prompt; review every line. 3) PII leakage — pasting sensitive data into external AI tools. Mitigation: use anonymised or synthetic data; use on-premise models (Snowflake Cortex) for sensitive work. 4) Over-reliance — analysts lose core SQL/Python skills. Mitigation: always understand what the AI generated; use AI to accelerate, not replace thinking."),
]
theory_block(ws_ai, r4+1, ai_theory)
col_w(ws_ai, [20, 30, 40, 34, 30])

# ── SAVE ─────────────────────────────────────────────────────────
out = r"C:\Users\sriva\OneDrive\Desktop\Github\Interview_Prep_Workbook_FIXED.xlsx"
wb.save(out)
print(f"Saved: {out}")
